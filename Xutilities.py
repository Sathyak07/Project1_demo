#importing openpyxl module
import openpyxl

#creating a predefined function for getting row_count
def get_row_count(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_row

#creating a predefined function for getting column_count
def get_column_count(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_column

#creating a predefined function for reading data from excel file
def read_data(filename, sheetname, rownum, columnnum):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value

#creating a predefined function for writing data to excel file
def write_data(filename, sheetname, rownum, columnnum, input_data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value=input_data
    workbook.save(filename)